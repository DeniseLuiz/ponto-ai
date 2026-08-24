import openpyxl
from pathlib import Path


def aggregate_and_export(results: list[str], job_id: int, storage_dir: str) -> str:
    """
    Recebe a lista de textos (um por chunk processado pelo Gemini, no formato
    TAB-separated) e monta um único arquivo .xlsx.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado"

    alerts = []

    for block in results:
        if not block:
            continue
        for line in block.strip().splitlines():
            line = line.strip()
            if not line:
                continue
            # Detecta linhas de alerta (não tabulares) e as separa
            if "\t" not in line:
                alerts.append(line)
                continue
            row = line.split("\t")
            ws.append(row)

    if alerts:
        alert_ws = wb.create_sheet("Alertas")
        for i, alert in enumerate(alerts, start=1):
            alert_ws.cell(row=i, column=1, value=alert)

    out_dir = Path(storage_dir) / "results"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"result_{job_id}.xlsx"
    wb.save(path)
    return str(path)
